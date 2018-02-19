#!/usr/bin/env python
# -*- coding=utf-8 -*-

import xlrd
import xlwt
from datetime import date,datetime

def read_exc():
    workbook = xlrd.open_workbook(r'/Users/Lix/Desktop/test.xlsx')
    print workbook.sheet_names()

    main_sheet = workbook.sheet_by_index(0)
    # sheet的名称，行数，列数
    print main_sheet.name, main_sheet.nrows,main_sheet.ncols

    # 获取整行和整列的值（数组）
    rows = main_sheet.row_values(2) # 获取第三行内容
    cols = main_sheet.col_values(2) # 获取第三列内容
    print rows
    print cols

    # 获取单元格内容
    print main_sheet.cell(1,1).value.encode('utf-8')
    print main_sheet.cell_value(1,1).encode('utf-8')
    print main_sheet.row(1)[1].value.encode('utf-8')

    # 获取单元格内容的数据类型
    print main_sheet.cell(1,1).ctype

    # 时间格式格式化
    date_value = xlrd.xldate_as_tuple(main_sheet.cell_value(1, 0), workbook.datemode)
    print date(*date_value[:3]).strftime('%Y/%m/%d')

def xlwings():
    import xlwings as xw
    #连接到excel
    workbook = xw.Book(r'path/myexcel.xlsx')#连接excel文件
    #连接到指定单元格
    data_range = workbook.sheets('Sheet1').range('A1')
    #写入数据
    data_range.value = [1,2,3]
    #保存
    workbook.save()

def xlswriter():
    import xlsxwriter as xw
    #新建excel
    workbook  = xw.Workbook('myexcel.xlsx')
    #新建工作薄
    worksheet = workbook.add_worksheet()
    #写入数据
    worksheet.wirte('A1',1)
    #关闭保存
    workbook.close()

def xlutils():
    """xlutils基本代码import xlrd 读取数据
    """

    import xlwt #写入数据
    import xlutils #操作excel
    #----xlrd库
    #打开excel文件
    workbook = xlrd.open_workbook('myexcel.xls')
    #获取表单
    worksheet = workbook.sheet_by_index(0)
    #读取数据
    data = worksheet.cell_value(0,0)
    #----xlwt库
    #新建excel
    wb = xlwt.Workbook()
    #添加工作薄
    sh = wb.add_sheet('Sheet1')
    #写入数据
    sh.write(0,0,'data')
    #保存文件
    wb.save('myexcel.xls')
    #----xlutils库
    #打开excel文件
    book = xlrd.open_workbook('myexcel.xls')
    #复制一份
    new_book = xlutils.copy(book)
    #拿到工作薄
    worksheet = new_book.getsheet(0)
    #写入数据
    worksheet.write(0,0,'new data')
    #保存
    new_book.save()

def openpyxl():
    """openpyxl基本代码 
    """
    import openpyxl
    # 新建文件
    workbook = openpyxl.Workbook() 
    # 写入文件
    sheet = workbook.activesheet['A1']='data'
    # 保存文件 
    workbook.save('test.xlsx')

def dataNitro():
    """DataNitro基本代码
    """
    #单一单元格赋值
    Cell('A1').value = 'data'
    #单元区域赋值
    CellRange('A1:B2').value = 'data'
    
def main():
    read_exc()

if __name__ == '__main__':
    main()